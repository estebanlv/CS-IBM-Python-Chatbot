import time
from array import *
from openpyxl import load_workbook #This loads a library that allows me to read/write excel files

wb = load_workbook(filename = 'wordsdatabase.xlsx') #Tell the program what file i'm using
linkdatabase = load_workbook(filename = 'sources.xlsx')
sheet = wb['main'] #Tells the program which sheet in the document i'm using
secsheet = wb['secondary']
total = wb['total']
linksheet = linkdatabase['main']
letters = ["A","B","C","D","E","F","G","H","I","K","J","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
topics = [["compute", 0],["containers", 0],["networking", 0],["storage", 0],["ai", 0],["analytics", 0],["databases", 0],["devtools", 0],["integration", 0],["iot", 0],["secandid", 0],["starterkits", 0],["webmobile", 0],["webapp", 0]]
linkcell = ["C2","C4","C6","C8","C10","C12","C14","C16","C18","C20","C22","C24","C26","C28"]
#These variable above is for cell values

def removepunctuation():
    output = open('dissectedmessage.txt', 'a') #opens file to append it.
    with open('message.txt', buffering = 20000000) as f: #opens the text file
        for line in f: #for every line in the file...
        #this code below is used to remove the punctuation from the text
            saveline = line.replace('.',' ')
            saveline = saveline.replace(',','')
            saveline = saveline.replace(':','')
            saveline = saveline.replace(';','')
            saveline = saveline.replace('!','')
            saveline = saveline.replace('?','')
            saveline = saveline.replace('"','')
            saveline = saveline.replace('#','')
            saveline = saveline.replace('-','')
            saveline = saveline.replace('_','')
            saveline = saveline.replace('=','')
            saveline = saveline.replace('+','')
            saveline = saveline.replace('@','')
            saveline = saveline.replace('(','')
            saveline = saveline.replace(')','')
            saveline = saveline.replace('[','')
            saveline = saveline.replace(']','')
            saveline = saveline.replace('*','')
            saveline = saveline.replace('%','')
            saveline = saveline.replace('£','')
            saveline = saveline.replace('$','')
            saveline = saveline.replace('  ',' ')
            saveline = saveline.replace('\n',' ')
            #saveline = saveline.rstrip() #removes all the newlines \n
            saveline = saveline.replace(' ','\n') #puts every word in a different line
            saveline = str.lower(saveline) #converts everything to lowercase
            output.write(saveline)
    output.close() #saves and closes the file

def joinandsplitspace(word_list):
    joined_text = ''.join(word_list) #joins the list to create a string
    #print(joined_text)
    list_of_words = joined_text.split()#Splits the string at every space
    lword_list = len(list_of_words) #Counts the words within the list
    #for x in range(lword_list): #for every word in this list...
        #print(list_of_words[x]) #prints all the words in the sentence
    return list_of_words;

def readfile():
    #topic = input("Enter the main topic of this text:")
    text_file = open("dissectedmessage.txt", "r") #opens the text file to read it
    word_list = text_file.read()
    print("read")
    return word_list;

def writemessagetofile(messagetext):
    message_file = open("message.txt", "w")
    message_file.write(messagetext)
    message_file.close()

def setrelevancetotopics(relevance, ystr):
    print("I work")
    no_of_topics_cell = letters[4] + ystr #gets the cell where the number of topics is
    no_of_topics  = int(sheet[no_of_topics_cell].value) #gets the value
    no_of_topics = no_of_topics + 5 #Represents the value of the column where the last topic should be found
    for z in range(4,no_of_topics): #for every topic do the following
        the_topic_cell = letters[z] + ystr #create the cell for the topic
        the_topic = sheet[the_topic_cell].value #gets the topic
        for i in range(0, 14): #for every possible topic do
            if the_topic == topics[i][0]: #if the database topic is the same as the topic in the array do
                topics[i][1] = topics[i][1] + relevance #add the relevance to the total

def relevanceextraction(list_of_words):
    word_list_length = len(list_of_words)
    print(word_list_length)
    for x in range(word_list_length): #for every word in the list do
        failed_attempts = 0 #variable used to count how many failed tries did it take to get the right word
        for y in range(2,10001): #repeat 10000 times. so every word is checked
            ystr = str(y) #transforms the value into a string in order to create the cell
            cell = letters[0] + ystr #creates the cell for the new word
            word_from_database = sheet[cell].value #saves the value from the cell inside a local variable
            word_from_sec_database = secsheet[cell].value #saves the value from the cell inside a local variable
            #print(word_from_database)
            if list_of_words[x] == word_from_database: #checks to see if the word is the same as the one in the database
                print("same")
                relevance_cell = letters[3] + ystr
                relevance = float(sheet[relevance_cell].value)
                setrelevancetotopics(relevance, ystr)
            elif list_of_words[x] == word_from_sec_database: #if the word is the same as the one in the secondary database
                print("same sec")
                relevance_cell = letters[3] + ystr
                relevance = float(secsheet[relevance_cell].value)
                setrelevancetotopics(relevance, ystr)
            else: #if the word is not in the database then...
                failed_attempts = failed_attempts + 1 #for every failed attempt, 1 is added to the variable
        if failed_attempts == 9999: #if every word in the main file failed then...
            print("nothing")

def topicalgorithm():
    no_of_topics = 13 #There are actually 14 topics but in the array the num 14 is represented as 13
    best_topic_score = 0
    best_topic = ""
    for x in range(0,13): #for every topic do...
        score = topics[x][1] #takes the score from the topic
        if score > best_topic_score: #if the current score is better than the best score do:
            best_topic_score = score #best score is updated with the new one
            best_topic = topics[x][0] #Saves the topic name with the best score
    if best_topic_score == 0:
        best_topic = "none"
    print(best_topic)
    return best_topic; #returns the best topic

def choosethelink(best_topic):
	for i in range(0,13): #for every topic do...
		if best_topic == topics[i][0]: #if the best topic is equal to the topic saved in the array
			link_cell = linkcell[i] #saves the cell number assigned to that topic
			link = str(linksheet[link_cell].value) #Saves the link that is stored in the database
	return link; #returns the link

def deletecontents():
    for x in range(0,13): #for every topic do...
        topics[x][1] = 0 #change the relevance to 0
    #this part below resets the dissectedmessage file in order to reuse it again
    f = open("dissectedmessage.txt", "w")
    f.close()

def start(messagetext):
    print("Lets start the search")
    writemessagetofile(messagetext)
    removepunctuation()
    word_list = readfile()
    list_of_words = joinandsplitspace(word_list)
    print(list_of_words)
    relevanceextraction(list_of_words)
    for x in range(0,14):
        #debug to see the relevances
        print(topics[x][0])
        print(topics[x][1])
    best_topic = topicalgorithm()
    if best_topic == "none":
        deletecontents()
        link = "Sorry but your request was unsuccesful, please try again"
        return(link)
    else:
        message = choosethelink(best_topic)
        deletecontents()
        print(message)
        return message; #sends the link to the main program
