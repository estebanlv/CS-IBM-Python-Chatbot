import time
from openpyxl import load_workbook #This loads a library that allows me to read/write excel files

start = time.time()

wb = load_workbook(filename = 'wordsdatabase.xlsx') #Tell the program what file i'm using
sheet = wb['main'] #Tells the program which sheet in the document i'm using
secsheet = wb['secondary']
total = wb['total']
letters = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
#These variable above is for cell values

def removepunctuation():
    output = open('trialfixed.txt', 'a') #opens file to append it.
    with open('trial.txt', buffering = 20000000) as f: #opens the text file
        for line in f: #for every line in the file...
        #this code below is used to remove the punctuation from the text
            saveline = line.replace('.',' ')
            saveline = saveline.replace(',','')
            saveline = saveline.replace(':','')
            saveline = saveline.replace(';','')
            saveline = saveline.replace('!','')
            saveline = saveline.replace('?','')
            saveline = saveline.replace('"','')
            saveline = saveline.replace('#','')
            saveline = saveline.replace('-','')
            saveline = saveline.replace('_','')
            saveline = saveline.replace('=','')
            saveline = saveline.replace('+','')
            saveline = saveline.replace('@','')
            saveline = saveline.replace('(','')
            saveline = saveline.replace(')','')
            saveline = saveline.replace('[','')
            saveline = saveline.replace(']','')
            saveline = saveline.replace('*','')
            saveline = saveline.replace('%','')
            saveline = saveline.replace('£','')
            saveline = saveline.replace('$','')
            saveline = saveline.replace('  ',' ')
            saveline = saveline.replace('\n',' ')
            #saveline = saveline.rstrip() #removes all the newlines \n
            saveline = saveline.replace(' ','\n') #puts every word in a different line
            saveline = str.lower(saveline) #converts everything to lowercase
            output.write(saveline)
    output.close() #saves and closes the file

def addtothetotal():
    #add 1 to the total amount of words recorded for the percentage stats
    total_cell = "B2" #This is the exact cell where this data has to be stored
    total_words_recorded = int(total[total_cell].value) #extracts the number from the cell
    total_words_recorded = total_words_recorded + 1 #adds one to that value
    total_words_recorded_str = str(total_words_recorded) #transforms it into a string
    total[total_cell].value = total_words_recorded_str #stores the number-string in the cell that it came from

def secwordsrecorded():
    #stores the amount of times words have been stored in the secondary database (reps included)
    total_recorded_sec_cell = "B4" #cell where the data is found
    total_words_recorded = int(total[total_recorded_sec_cell].value) #extracts the number from the cell
    total_words_recorded = total_words_recorded + 1 #adds one to that value
    total_words_recorded_str = str(total_words_recorded) #transforms it into a string
    total[total_recorded_sec_cell].value = total_words_recorded_str #stores the number-string in the cell that it came from

def dostats(ystr):
    #number of apparitions statistics
    cell = letters[1] + ystr #creates the cell e.g. [B2]
    apps_number = int(sheet[cell].value) #saves the number in a integer variable
    apps_number =  apps_number + 1 #adds 1 to the variable
    apps_num_str = str(apps_number) #stores it as a string
    sheet[cell].value = apps_num_str #stores the updated number in the same cell it came from e.g. [B2]
    addtothetotal()

def dosecstats(ystr):
    #number of apparitions statistics
    cell = letters[1] + ystr #creates the cell e.g. [B2]
    apps_number = int(secsheet[cell].value) #saves the number in a integer variable
    apps_number =  apps_number + 1 #adds 1 to the variable
    apps_num_str = str(apps_number) #stores it as a string
    secsheet[cell].value = apps_num_str #stores the updated number in the same cell it came from e.g. [B2]
    secwordsrecorded()

def savetheword(word, topic):
    #saves the unknown words into the secondary database
    tot_words_insec_cell = "B3" #where the total num of words in the secondary database is stored
    num_of_tot_words_insec = int(total[tot_words_insec_cell].value) #gets the number of total words in sec. database
    row_number = num_of_tot_words_insec + 2 #next row number availeable (no data inside of it)
    row_number_str = str(row_number) #converts it into a string
    word_save_cell = letters[0] + row_number_str #makes the cell where there isnt any data in
    secsheet[word_save_cell].value = word #stores the word on that cell
    num_of_tot_words_insec_str = str(num_of_tot_words_insec + 1) #converts it into a string
    total[tot_words_insec_cell].value = num_of_tot_words_insec_str #stores the new amount of words in the sec database
    no_apps_cell = letters[1] + row_number_str #get the cell where the apps are stored
    no_of_apps_int = int(secsheet[no_apps_cell].value) #gets the value and transforms it into an integer
    no_of_apps_str = str(no_of_apps_int + 1) #add one apparition to the number
    secsheet[no_apps_cell].value =  no_of_apps_str #saves it back to the cell as a string
    #saves the topic_cell
    empty_topic_cell = letters[5] + row_number_str #creates the cell where the first topic should be at
    secsheet[empty_topic_cell].value = topic #stores the topic in the empty cell
    number_of_topics = 1 #sets the number of topics to one
    number_of_topics_str = str(number_of_topics) #stores it as a string
    topic_no_cell = letters[4] + row_number_str
    secsheet[topic_no_cell].value = number_of_topics_str #stores the value at the respective cell
    secwordsrecorded()

def joinandsplitspace(word_list):
    joined_text = ''.join(word_list) #joins the list to create a string
    #print(joined_text)
    list_of_words = joined_text.split()#Splits the string at every space
    lword_list = len(list_of_words) #Counts the words within the list
    #for x in range(lword_list): #for every word in this list...
        #print(list_of_words[x]) #prints all the words in the sentence
    return list_of_words;

def dotopics(ystr, topic):
    notequal = 0
    #selects the next empty cell
    topic_no_cell = letters[4] + ystr #creates a cell that contains the number of topics of a word
    number_of_topics = int(sheet[topic_no_cell].value) #extracts the number of topics of a word
    if number_of_topics == 0: #if there isnt any topics...
        num_letter = number_of_topics + 5 #first cell column letter [F]
        empty_topic_cell = letters[num_letter] + ystr #creates the cell where the first topic should be at
        sheet[empty_topic_cell].value = topic #stores the topic in the empty cell
        number_of_topics = number_of_topics + 1 #increases the number of topics by one
        number_of_topics_str = str(number_of_topics) #stores it as a string
        sheet[topic_no_cell].value = number_of_topics_str #stores the value at the respective cell
    elif number_of_topics != 0: #if the number of topics isnt 0 then...
        num_letter = number_of_topics + 4
        for y in range(number_of_topics): #for every topic do...
            if num_letter > 25:
                num_letter = 25
            topic_cell = letters[num_letter] + ystr #create cell where the topic is stored
            topic_database = str(sheet[topic_cell].value) #extracts the string within the cell
            if topic == topic_database: #if the topic is the same do absolutely nothing
                print("same")
            else: #if it isnt the same...
                notequal = notequal + 1
                num_letter = num_letter + 1
    if notequal == number_of_topics: #if the topic is not within the topic list already do...
        num_letter = number_of_topics + 5 #checks the last empty column
        empty_topic_cell = letters[num_letter] + ystr #creates the empty cell
        sheet[empty_topic_cell].value = topic #saves the topic on the cell
        number_of_topics = number_of_topics + 1 #increases the number of topics by one
        number_of_topics_str = str(number_of_topics) #stores it as a string
        sheet[topic_no_cell].value = number_of_topics_str #stores the value at the respective cell

def dosectopics(ystr, topic):
    notequal = 0
    #selects the next empty cell
    topic_no_cell = letters[4] + ystr #creates a cell that contains the number of topics of a word
    number_of_topics = int(secsheet[topic_no_cell].value) #extracts the number of topics of a word
    if number_of_topics == 0: #if there isnt any topics...
        num_letter = number_of_topics + 5 #first cell column letter [F]
        empty_topic_cell = letters[num_letter] + ystr #creates the cell where the first topic should be at
        secsheet[empty_topic_cell].value = topic #stores the topic in the empty cell
        number_of_topics = number_of_topics + 1 #increases the number of topics by one
        number_of_topics_str = str(number_of_topics) #stores it as a string
        secsheet[topic_no_cell].value = number_of_topics_str #stores the value at the respective cell
    elif number_of_topics != 0: #if the number of topics isnt 0 then...
        num_letter = number_of_topics + 4
        for y in range(number_of_topics): #for every topic do...
            if num_letter > 25:
                num_letter = 25
            topic_cell = letters[num_letter] + ystr #create cell where the topic is stored
            topic_database = str(secsheet[topic_cell].value) #extracts the string within the cell
            if topic == topic_database: #if the topic is the same do absolutely nothing
                print("same")
            else: #if it isnt the same...
                notequal = notequal + 1
                num_letter = num_letter + 1
    if notequal == number_of_topics: #if the topic is not within the topic list already do...
        num_letter = number_of_topics + 5 #checks the last empty column
        empty_topic_cell = letters[num_letter] + ystr #creates the empty cell
        secsheet[empty_topic_cell].value = topic #saves the topic on the cell
        number_of_topics = number_of_topics + 1 #increases the number of topics by one
        number_of_topics_str = str(number_of_topics) #stores it as a string
        secsheet[topic_no_cell].value = number_of_topics_str #stores the value at the respective cell

def wordcheck():
    topic = input("Enter the main topic of this text:")
    text_file = open("trialfixed.txt", "r") #opens the text file to read it
    word_list = text_file.read()
    print("read")
    words_list = joinandsplitspace(word_list)
    word_list_length = len(words_list)
    print(word_list_length)
    #timex = time.time()
    for x in range(word_list_length): #for every word in the list do
        #print(words_list[x])
        #timex = time.time()
        #print (x, '/', word_list_length)
        failed_attempts = 0 #variable used to count how many failed tries did it take to get the right word
        for y in range(2,10001): #repeat 10000 times. so every word is checked
            ystr = str(y) #transforms the value into a string in order to create the cell
            cell = letters[0] + ystr #creates the cell for the new word
            word_from_database = sheet[cell].value #saves the value from the cell inside a local variable
            word_from_sec_database = secsheet[cell].value #saves the value from the cell inside a local variable
            #print(word_from_database)
            if words_list[x] == word_from_database: #checks to see if the word is the same as the one in the database
                #print("same")
                dostats(ystr) #if it is, then do the statistics part
                dotopics(ystr, topic)
            elif words_list[x] == word_from_sec_database: #if the word is the same as the one in the secondary database
                #print("same sec")
                dosecstats(ystr) #if it is, then do the sec statistics part
                dosectopics(ystr, topic)
            else: #if the word is not in the database then...
                failed_attempts = failed_attempts + 1 #for every failed attempt, 1 is added to the variable
        #if x == 100:
            #endx = time.time()
            #print(endx-timex)
        if failed_attempts == 9999: #if every word in the main file failed then...
            #print("not same")
            word = words_list[x] #saves the word to pass it into the function
            savetheword(word, topic) #calls the savetheword function
        #endx = time.time()
        #print(endx-timex)

def calculatepercentage():
    for y in range(2,10001):
        ystr = str(y) #transforms the value into a string in order to create the cell
        word_app_cell = letters[1] + ystr #creates the cell for the new word
        total_cell = "B2" #cell where the total is stored at
        word_no_of_apps = int(sheet[word_app_cell].value) #stores the apps from a single word
        total_no_of_apps = int(total[total_cell].value) #stores the apps in total (all words)
        word_percentage = (word_no_of_apps / total_no_of_apps) * 100 #percentage calculation
        word_percentage_str = str(word_percentage) #transform into a string to store in database
        percentage_cell = letters[2] + ystr #creates the cell for the percentage
        sheet[percentage_cell].value = word_percentage_str #stores the percentage value

def calculatesecpercentage():
    row_count = secsheet.max_row #checks how many rows are in the secondary database
    for y in range(2,row_count):
        ystr = str(y) #transforms the value into a string in order to create the cell
        word_app_cell = letters[1] + ystr #creates the cell for the new word
        total_cell = "B4" #cell where the total is stored at
        word_no_of_apps = int(secsheet[word_app_cell].value) #stores the apps from a single word
        total_no_of_apps = int(total[total_cell].value) #stores the apps in total (all words)
        word_percentage = (word_no_of_apps / total_no_of_apps) * 100 #percentage calculation
        word_percentage_str = str(word_percentage) #transform into a string to store in database
        percentage_cell = letters[2] + ystr #creates the cell for the percentage
        secsheet[percentage_cell].value = word_percentage_str #stores the percentage value

print("This program will read anything on a text file and save the words into a database")
removepunctuation()
wordcheck()
calculatepercentage()
calculatesecpercentage()
#wb.save("wordsdatabase.xlsx")

end = time.time()
print(end - start) #time taken for the program to run
