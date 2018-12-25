from openpyxl import load_workbook #This loads a library that allows me to read/write excel files
wb = load_workbook(filename = 'wordsdatabase.xlsx') #Tell the program what file i'm using
sheet = wb['main'] #Tells the program which sheet in the document i'm using
secsheet = wb['secondary']
total = wb['total']
letters = ["A","B","C","D","E","F","G","H","I","K","J","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
#These variable above is for cell values

def removepunctuation():
    text_file = open("texttoread.txt", "r") #opens the text file to read it
    lines = text_file.read() #reads every line in the file and stores it in a list
    lines = lines.lower() #converts the whole string into lowercase
    character_list = list(lines) #separates the string into characters
    length_character_list = len(character_list) #counts the amount of characters in a list
    for i in range(length_character_list): #loop to go through all the characters in the list
        #This code deletes all the punctuation characters in the list
        if "," in character_list:
            character_list.remove(",")
        elif "." in character_list:
            character_list.remove(".")
        elif ":" in character_list:
            character_list.remove(":")
        elif "!" in character_list:
            character_list.remove("!")
        elif "?" in character_list:
            character_list.remove("?")
        elif "(" in character_list:
            character_list.remove("(")
        elif ")" in character_list:
            character_list.remove(")")
        elif "/" in character_list:
            character_list.remove("/")
        elif "'" in character_list:
            character_list.remove("'")
        elif "[" in character_list:
            character_list.remove("[")
        elif "]" in character_list:
            character_list.remove("]")
        elif "=" in character_list:
            character_list.remove("=")
        elif "+" in character_list:
            character_list.remove("+")
        elif "£" in character_list:
            character_list.remove("£")
        elif "$" in character_list:
            character_list.remove("$")
        elif "*" in character_list:
            character_list.remove("*")
        elif ";" in character_list:
            character_list.remove(";")
    print (character_list)
    text_file.close() #closes the file
    return character_list; #returns the list of character for the next section

def joinandsplitspace(character_list):
    joined_text = ''.join(character_list) #joins the list to create a string
    print(joined_text)
    list_of_words = joined_text.split()#Splits the string at every space
    lword_list = len(list_of_words) #Counts the words within the list
    for x in range(lword_list): #for every word in this list...
        print(list_of_words[x]) #prints all the words in the sentence
    return list_of_words;

def addtothetotal():
    #add 1 to the total amount of words recorded for the percentage stats
    total_cell = "B2" #This is the exact cell where this data has to be stored
    total_words_recorded = int(total[total_cell].value) #extracts the number from the cell
    total_words_recorded = total_words_recorded + 1 #adds one to that value
    total_words_recorded_str = str(total_words_recorded) #transforms it into a string
    total[total_cell].value = total_words_recorded_str #stores the number-string in the cell that it came from

def secwordsrecorded():
    #stores the amount of times words have been stored in the secondary database (reps included)
    total_recorded_sec_cell = "B4" #cell where the data is found
    total_words_recorded = int(total[total_recorded_sec_cell].value) #extracts the number from the cell
    total_words_recorded = total_words_recorded + 1 #adds one to that value
    total_words_recorded_str = str(total_words_recorded) #transforms it into a string
    total[total_recorded_sec_cell].value = total_words_recorded_str #stores the number-string in the cell that it came from

def dostats(ystr):
    #number of apparitions statistics
    cell = letters[1] + ystr #creates the cell e.g. [B2]
    apps_number = int(sheet[cell].value) #saves the number in a integer variable
    apps_number =  apps_number + 1 #adds 1 to the variable
    apps_num_str = str(apps_number) #stores it as a string
    sheet[cell].value = apps_num_str #stores the updated number in the same cell it came from e.g. [B2]
    addtothetotal()

def dosecstats(ystr):
    #number of apparitions statistics
    cell = letters[1] + ystr #creates the cell e.g. [B2]
    apps_number = int(secsheet[cell].value) #saves the number in a integer variable
    apps_number =  apps_number + 1 #adds 1 to the variable
    apps_num_str = str(apps_number) #stores it as a string
    secsheet[cell].value = apps_num_str #stores the updated number in the same cell it came from e.g. [B2]
    secwordsrecorded()

def savetheword(word):
    #saves the unknown words into the secondary database
    tot_words_insec_cell = "B3" #where the total num of words in the secondary database is stored
    num_of_tot_words_insec = int(total[tot_words_insec_cell].value) #gets the number of total words in sec. database
    row_number = num_of_tot_words_insec + 2 #next row number availeable (no data inside of it)
    row_number_str = str(row_number) #converts it into a string
    word_save_cell = letters[0] + row_number_str #makes the cell where there isnt any data in
    secsheet[word_save_cell].value = word #stores the word on that cell
    num_of_tot_words_insec_str = str(num_of_tot_words_insec + 1) #converts it into a string
    total[tot_words_insec_cell].value = num_of_tot_words_insec_str #stores the new amount of words in the sec database
    no_apps_cell = letters[1] + row_number_str #get the cell where the apps are stored
    no_of_apps_int = int(secsheet[no_apps_cell].value) #gets the value and transforms it into an integer
    no_of_apps_str = str(no_of_apps_int + 1) #add one apparition to the number
    secsheet[no_apps_cell].value =  no_of_apps_str #saves it back to the cell as a string
    secwordsrecorded()

def wordcheck(words_list):
    word_list_length = len(words_list) #checks the number of words in the list
    print(word_list_length)
    for x in range(word_list_length): #for evry word in the list do
        print(words_list[x])
        failed_attempts = 0 #variable used to count how many failed tries did it take to get the right word
        for y in range(2,10001): #repeat 10000 times. so every word is checked
            ystr = str(y) #transforms the value into a string in order to create the cell
            cell = letters[0] + ystr #creates the cell for the new word
            word_from_database = sheet[cell].value #saves the value from the cell inside a local variable
            word_from_sec_database = secsheet[cell].value #saves the value from the cell inside a local variable
            #print(word_from_database)
            if words_list[x] == word_from_database: #checks to see if the word is the same as the one in the database
                print("same")
                dostats(ystr) #if it is, then do the statistics part
            elif words_list[x] == word_from_sec_database: #if the word is the same as the one in the secondary database
                print("same sec")
                dosecstats(ystr) #if it is, then do the sec statistics part
            else: #if the word is not in the database then...
                failed_attempts = failed_attempts + 1 #for every failed attempt, 1 is added to the variable

        if failed_attempts == 9999: #if every word in the main file failed then...
            print("not same")
            word = words_list[x] #saves the word to pass it into the function
            savetheword(word) #calls the savetheword function

def calculatepercentage():
    for y in range(2,10001):
        ystr = str(y) #transforms the value into a string in order to create the cell
        word_app_cell = letters[1] + ystr #creates the cell for the new word
        total_cell = "B2" #cell where the total is stored at
        word_no_of_apps = int(sheet[word_app_cell].value) #stores the apps from a single word
        total_no_of_apps = int(total[total_cell].value) #stores the apps in total (all words)
        word_percentage = (word_no_of_apps / total_no_of_apps) * 100 #percentage calculation
        word_percentage_str = str(word_percentage) #transform into a string to store in database
        percentage_cell = letters[2] + ystr #creates the cell for the percentage
        sheet[percentage_cell].value = word_percentage_str #stores the percentage value

def calculatesecpercentage():
    row_count = secsheet.max_row #checks how many rows are in the secondary database
    for y in range(2,row_count):
        ystr = str(y) #transforms the value into a string in order to create the cell
        word_app_cell = letters[1] + ystr #creates the cell for the new word
        total_cell = "B4" #cell where the total is stored at
        word_no_of_apps = int(secsheet[word_app_cell].value) #stores the apps from a single word
        total_no_of_apps = int(total[total_cell].value) #stores the apps in total (all words)
        word_percentage = (word_no_of_apps / total_no_of_apps) * 100 #percentage calculation
        word_percentage_str = str(word_percentage) #transform into a string to store in database
        percentage_cell = letters[2] + ystr #creates the cell for the percentage
        secsheet[percentage_cell].value = word_percentage_str #stores the percentage value

print("This program will read anything on a text file and save the words into a database")
char_list = removepunctuation() #saves the list from the other def in another list
word_list = joinandsplitspace(char_list) #saves the word list from the procedure into another global variable
wordcheck(word_list) #calls the wordcheck function to run the words through a database
calculatepercentage()
calculatesecpercentage()
wb.save("wordsdatabase.xlsx")



#row_count = sheet.max_row
#print (row_count)

#text_file = open("texttoread.txt", "r") #opens the text file to read it
#lines = text_file.readlines() #reads every line in the file and stores it in a list
#for x in lines: #for every line in the list...
    #print (x) #prints all the lines
#text_file.close() #closes the file
